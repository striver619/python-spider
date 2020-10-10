
import MySQLdb
from openpyxl import Workbook,load_workbook

class ExcelUtil(object):

    def __init__(self):
        """
          导出数据到Excel
        """
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws_clean_data = self.wb.create_sheet("清洗数据")
        self.ws_clean_data.append(['姓名', '年龄', '性别', '身份证'])

    def do_add_data(self):
        """
        添加，保存数据
        """
        self.ws_clean_data["F6"] = 66
        for row in self.ws_clean_data['A1:E5']:
            for cell in row:
                cell.value = "Test"
        self.wb.save('./test.xlsx')


    def read_xls(self):
        """
         读取excel数据，导入到sql数据库
        :return:
        """
        ws = load_workbook('./template.xlsx')
        names = ws.get_sheet_names()
        print(names)

        conn = self.get_conn()
        wb = ws.active
        wb = ws[names[0]]
        for (i, row) in enumerate(wb.rows):
            if i < 2:
                continue
            year = wb['A{0}'.format(i + 1)].value
            max = wb['B{0}'.format(i + 1)].value
            avg = wb['C{0}'.format(i + 1)].value
            print(year)
            if year is None:
                continue
            cursor = conn.cursor()
            sql = 'INSERT INTO `house`(`id`, `name`, `price`) VALUES({year}, {max}, {avg})'.format(
                year=year, max=max, avg=avg)
            print(sql)
            cursor.execute(sql)
            conn.autocommit(True)

    def export_xls(self):
        """ 从mysql数据库导出数据到excel """
        # 获取数据库的连接
        conn = self.get_conn()
        cursor = conn.cursor()
        # 准备查询语句 (如果数据量大，需要借助于分页查询)
        sql = 'SELECT * FROM `house`'
        # 查询数据
        cursor.execute(sql)
        rows = cursor.fetchall()

        # 循环写入到excel
        wb = Workbook()
        ws = wb.active
        for (i, row) in enumerate(rows):
            print(row)
            (ws['A{0}'.format(i + 1)],
             ws['B{0}'.format(i + 1)],
             ws['C{0}'.format(i + 1)]) = row

        # 保存excel
        wb.save('./export.xlsx')


    def get_conn(self):
        """ 获取mysql 的连接 """
        conn = MySQLdb.connect('localhost', 'root', 'root', 'test', charset="utf8", use_unicode=True)
        return conn


if __name__ == '__main__':
    client = ExcelUtil()
    # client.do_add_data()
    client.read_xls()
    # client.export_xls()
